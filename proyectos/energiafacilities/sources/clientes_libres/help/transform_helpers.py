from __future__ import annotations
import logging
import re
from pathlib import Path
from datetime import datetime, date
import pandas as pd
from openpyxl import load_workbook
import os
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURACIÓN INTERNA / ESTÁNDARES
# ============================================================
ZERO_WHEN_EMPTY = True
DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]
CELL_RE = re.compile(r"^[A-Z]+[0-9]+$")
SPANISH_MONTHS = {
    "enero": 1, "ene": 1, "febrero": 2, "feb": 2, "marzo": 3, "mar": 3, "abril": 4, "abr": 4,
    "mayo": 5, "may": 5, "junio": 6, "jun": 6, "julio": 7, "jul": 7, "agosto": 8, "ago": 8,
    "setiembre": 9, "septiembre": 9, "sep": 9, "set": 9, "octubre": 10, "oct": 10,
    "noviembre": 11, "nov": 11, "diciembre": 12, "dic": 12
}

# ============================================================
# UTILIDADES INTERNAS
# ============================================================

def _parse_number(s: str) -> float:
    """Convierte cadenas numéricas con comas/puntos a float normalizado."""
    s = str(s).strip()
    if "," in s and s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    return float(s)


def _cast_value(val, type_spec: str, zero_when_empty: bool, date_formats: list[str]):
    """Convierte valores al tipo definido en mapping."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return 0 if type_spec in ("int", "decimal") and zero_when_empty else None

    try:
        if type_spec == "date":
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            if isinstance(val, str):
                for fmt in date_formats:
                    try:
                        return datetime.strptime(val.strip(), fmt).date()
                    except ValueError:
                        continue
            return None

        if type_spec == "int":
            return int(float(_parse_number(val)))

        if type_spec == "decimal":
            return float(_parse_number(val))

        return str(val).strip()

    except Exception:
        if type_spec in ("int", "decimal") and zero_when_empty:
            return 0
        return None


def _yyyymm_from_periodo(text: str | None) -> str | None:
    """
    Extrae un periodo en formato YYYYMM desde una cadena.
    Ejemplos válidos:
      - "Setiembre 2024"
      - "2024/09", "2024-09", "09/2024", "09-2024"
      - "202409"
    """
    if not text:
        return None

    s = str(text).strip().lower()

    # Patrones comunes con separadores: "2024-09" o "09/2024"
    match = re.search(r"(?:(20\d{2})[/-](0?[1-9]|1[0-2]))|((0?[1-9]|1[0-2])[/-](20\d{2}))", s)
    if match:
        y1, m1, _, m2, y2 = match.groups()
        if y1 and m1:
            return f"{int(y1)}{int(m1):02d}"
        if y2 and m2:
            return f"{int(y2)}{int(m2):02d}"

    # Patrones con nombre del mes
    ymatch = re.search(r"(20\d{2})", s)
    if ymatch:
        y = int(ymatch.group(1))
        for name, num in SPANISH_MONTHS.items():
            if name in s:
                return f"{y}{num:02d}"

    # Formato compacto tipo "202409"
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", s)
    if match:
        return f"{int(match.group(1))}{int(match.group(2)):02d}"

    return None


def _cod_to_str(cod) -> str | None:
    """Normaliza código de suministro a cadena limpia."""
    if cod is None:
        return None
    try:
        logger.debug(f"Convirtiendo código de suministro: {cod}")
        return str(int(float(cod)))
    except Exception:
        s = re.sub(r"\D", "", str(cod))
        return s if s else None
    


def _leer_registro(ws, mapping: dict) -> dict:
    """Lee un registro desde una hoja Excel aplicando el mapping definido."""
    registro = {}
    for campo, cfg in mapping.items():
        tipo = cfg.get("type", "str").lower()
        celda = cfg.get("cell")
        const = cfg.get("const")
        val = ws[celda].value if celda else const
        registro[campo] = _cast_value(val, tipo, ZERO_WHEN_EMPTY, DATE_FORMATS)
    logger.debug(f"Registro leído: {registro}")
    # Campos derivados
    registro["periodo"] = _yyyymm_from_periodo(registro.get("per_consumo"))
    registro["cod_suministro"] = _cod_to_str(registro.get("cod_suministro"))
    registro["num_recibo"] = (
        f"{registro['periodo']}{registro['cod_suministro']}"
        if registro["periodo"] and registro["cod_suministro"]
        else None
    )
    return registro


def _procesar_excel(path_xlsx: Path, mapping: dict, sheet_names: list[str]) -> pd.DataFrame:
    """Procesa las hojas indicadas del archivo Excel según el mapping."""
    wb = load_workbook(filename=path_xlsx, data_only=True, read_only=True)
    try:
        logger.debug(f"Hojas disponibles en {path_xlsx.name}")
        disponibles = set(wb.sheetnames)
        registros = []

        for nombre in sheet_names:
            if nombre not in disponibles:
                logger.warning(f"La hoja '{nombre}' no existe en {path_xlsx.name}.")
                continue
            ws = wb[nombre]
            registro = _leer_registro(ws, mapping)
            registro["hoja"] = nombre
            registros.append(registro)
        logger.debug(f"Hojas procesadas: {len(registros)} de {len(sheet_names)} solicitadas.")
        if not registros:
            logger.warning("No se generaron registros: todas las hojas fueron omitidas o vacías.")
            return pd.DataFrame(columns=list(mapping.keys()) + ["hoja", "num_recibo"])

        df = pd.DataFrame(registros)
        logger.debug(f"Procesamiento completado: {len(df)} registros generados.")
        return df
    finally:
        wb.close()

# ============================================================
# FUNCIÓN PRINCIPAL (solo transformación)
# ============================================================

def ejecutar_transformacion(config_transform: dict, mapeo_campos: dict, filepath:str=None, save: bool=False, newdestinationoptional: str =None) -> pd.DataFrame | str:
    """
    Lee un archivo Excel local y devuelve un DataFrame transformado.
    No guarda archivos, solo transforma datos en memoria.

    Parámetros:
      config_transform: {"local_destination_dir": "...", "specific_filename": "...", local_dir: "..."}
      mapeo_campos: {"sheet_names": [...], "mapping": {...}}

    Retorna:
      pandas.DataFrame con los datos transformados.
    """
    try:
        if(filepath):
            input_path =Path(filepath)
        else:
            local_dir = Path(config_transform.get("local_dir", ""))
            filename = config_transform.get("specific_filename")
            input_path = local_dir / filename
        mapping = mapeo_campos.get("mapping")
        sheet_names = mapeo_campos.get("sheet_names", [])

        # ===== Validación =====
        if not input_path:
            logger.error("Faltan parámetros de path de archivo en la configuración.")
            raise ValueError("Faltan parámetros de path de archivo en la configuración")

        if not mapping or not isinstance(mapping, dict):
            logger.error("El objeto 'mapping' es requerido en mapeo_campos.")
            raise ValueError("El objeto 'mapping' es requerido en mapeo_campos")

        if not sheet_names:
            logger.error("Debe especificarse al menos una hoja en 'sheet_names'.")
            raise ValueError("Debe especificarse al menos una hoja en 'sheet_names'")

    
        if not input_path.exists():
            logger.error(f"No se encontró el archivo Excel: {input_path}")
            raise FileNotFoundError(f"No se encontró el archivo Excel: {input_path}")

        # ===== Transformación =====
        logger.debug(f"Iniciando transformación de archivo: {input_path}")
        df = _procesar_excel(input_path, mapping, sheet_names)
        # quitamos la columna periodo y la columna hoja porque no son necesarias en el resultado final
        if 'hoja' in df.columns:
            df = df.drop(columns=['hoja'])
        if 'periodo' in df.columns:
            df = df.drop(columns=['periodo'])

        if(save):
            filename = config_transform.get("local_destination_dir") or newdestinationoptional
            if not filename:
                logger.error("No se especificó un directorio de destino para guardar")
                raise
            try:
                pathcrear=filename.rsplit('/',1)[0]
                os.makedirs(pathcrear, exist_ok=True)
                logger.debug(f"Directorio {pathcrear} creado/existente.")
            except Exception as e:
                logger.error(f"No se pudo crear el directorio {pathcrear}: {e}")
                raise
            df.to_excel(filename, index=False)
            logger.debug(f"Archivo transformado guardado en: {filename}")
            return filename
        return df

    except Exception:
        logger.exception("Error crítico durante la transformación del Excel.")
        raise  # Raise vacío: conserva stacktrace en logs y corta flujo
