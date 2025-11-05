"""
Tests para validar correcciones de FASE 2 - Resource Leaks
Bugs: #1, #2, #5, #7
"""
import pytest
import sys
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch, call
import tempfile
import os

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "proyectos" / "energiafacilities"))
from core.base_stractor import BaseExtractorSFTP
from core.base_postgress import PostgresConnector
from core.base_run_sp import run_sp


class TestBug1_ConexionesSFTPSinCerrar:
    """
    BUG #1: Conexiones SFTP sin cerrar
    Verificar que sftp.close() se llama siempre, incluso con excepciones
    """

    def test_listar_archivos_cierra_conexion_en_exito(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que listar_archivos cierra conexión en caso exitoso"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client):
                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                # Ejecutar operación exitosa
                result = extractor.listar_archivos()

                # Verificar que se cerró la conexión
                assert mock_sftp_client.close.called
                assert mock_sftp_client.close.call_count == 1

    def test_listar_archivos_cierra_conexion_en_error(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que listar_archivos cierra conexión incluso si falla"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            # Hacer que listdir lance excepción
            mock_sftp_client.listdir.side_effect = Exception("SFTP Error")

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client):
                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                # Debe lanzar excepción pero cerrar conexión
                with pytest.raises(Exception, match="SFTP Error"):
                    extractor.listar_archivos()

                # Verificar que se cerró la conexión en el bloque finally
                assert mock_sftp_client.close.called

    def test_listar_archivos_atributos_cierra_conexion_en_exito(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que listar_archivos_atributos cierra conexión"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            mock_attr = MagicMock()
            mock_attr.filename = 'test.xlsx'
            mock_attr.st_mtime = 1234567890
            mock_sftp_client.listdir_attr.return_value = [mock_attr]

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client):
                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                result = extractor.listar_archivos_atributos()

                # Verificar que se cerró
                assert mock_sftp_client.close.called

    def test_listar_archivos_atributos_cierra_conexion_en_error(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que listar_archivos_atributos cierra conexión en error"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            mock_sftp_client.listdir_attr.side_effect = Exception("Error listing")

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client):
                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                with pytest.raises(Exception):
                    extractor.listar_archivos_atributos()

                assert mock_sftp_client.close.called

    def test_extract_cierra_conexion_en_exito(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que extract cierra conexión en caso exitoso"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client), \
                 patch('core.base_stractor.os.makedirs'):

                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                result = extractor.extract(specific_file='test.xlsx')

                # Verificar que se cerró
                assert mock_sftp_client.close.called
                assert result['status'] == 'success'

    def test_extract_cierra_conexion_en_error(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que extract cierra conexión incluso si falla"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            mock_sftp_client.get.side_effect = Exception("Download failed")

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client), \
                 patch('core.base_stractor.os.makedirs'):

                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                with pytest.raises(Exception, match="Download failed"):
                    extractor.extract(specific_file='test.xlsx')

                # Verificar que se cerró en finally
                assert mock_sftp_client.close.called


class TestBug2_TransportSFTPSinCerrar:
    """
    BUG #2: Transport SFTP sin cerrar
    Verificar que transport.close() se llama si falla la creación del client
    """

    def test_conectar_sftp_cierra_transport_en_error_connect(self, sftp_config_connect, sftp_config_paths):
        """Test que cierra transport si falla transport.connect()"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport.connect.side_effect = Exception("Connection failed")
            mock_transport_class.return_value = mock_transport

            extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

            with pytest.raises(ConnectionError):
                extractor.conectar_sftp()

            # Verificar que se cerró el transport
            assert mock_transport.close.called

    def test_conectar_sftp_cierra_transport_en_error_sftp_client(self, sftp_config_connect, sftp_config_paths):
        """Test que cierra transport si falla SFTPClient.from_transport()"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class, \
             patch('core.base_stractor.paramiko.SFTPClient.from_transport') as mock_from_transport:

            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport
            mock_from_transport.side_effect = Exception("SFTP client failed")

            extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

            with pytest.raises(ConnectionError):
                extractor.conectar_sftp()

            # Verificar que se cerró el transport
            assert mock_transport.close.called

    def test_conectar_sftp_no_cierra_transport_en_exito(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que NO cierra transport si tiene éxito (el client lo usa)"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class, \
             patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client):

            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

            sftp = extractor.conectar_sftp()

            # NO debe cerrar el transport en caso exitoso (lo usa el client)
            assert not mock_transport.close.called
            assert sftp == mock_sftp_client


class TestBug5_ConexionPostgreSQLSinCerrar:
    """
    BUG #5: Conexión PostgreSQL sin cerrar
    Verificar que se usa context manager y conexión se cierra
    """

    def test_run_sp_cierra_conexion_con_context_manager(self, postgres_config):
        """Test que run_sp usa context manager y cierra conexión"""
        with patch('core.base_run_sp.load_config') as mock_load_config, \
             patch('core.base_run_sp.PostgresConnector') as mock_pg_class:

            mock_load_config.return_value = {
                'postgress': postgres_config,
                'test_config': {'sp_carga': 'public.test_sp'}
            }

            mock_pg_instance = MagicMock()
            mock_pg_instance.__enter__ = Mock(return_value=mock_pg_instance)
            mock_pg_instance.__exit__ = Mock(return_value=False)
            mock_pg_instance.ejecutar.return_value = Mock(
                __getitem__=lambda self, key: Mock(values=['OK'])
            )

            mock_pg_class.return_value = mock_pg_instance

            # Ejecutar
            run_sp('test_config')

            # Verificar que se usó context manager
            assert mock_pg_instance.__enter__.called
            assert mock_pg_instance.__exit__.called

    def test_run_sp_cierra_conexion_incluso_con_error(self, postgres_config):
        """Test que run_sp cierra conexión incluso si hay error"""
        with patch('core.base_run_sp.load_config') as mock_load_config, \
             patch('core.base_run_sp.PostgresConnector') as mock_pg_class:

            mock_load_config.return_value = {
                'postgress': postgres_config,
                'test_config': {'sp_carga': 'public.test_sp'}
            }

            mock_pg_instance = MagicMock()
            mock_pg_instance.__enter__ = Mock(return_value=mock_pg_instance)
            mock_pg_instance.__exit__ = Mock(return_value=False)
            mock_pg_instance.ejecutar.side_effect = Exception("SP failed")

            mock_pg_class.return_value = mock_pg_instance

            # Debe lanzar excepción
            with pytest.raises(Exception, match="SP failed"):
                run_sp('test_config')

            # Pero debe haber llamado __exit__ (cierre del context manager)
            assert mock_pg_instance.__enter__.called
            assert mock_pg_instance.__exit__.called


class TestBug7_WorkbookExcelSinCerrar:
    """
    BUG #7: Workbook Excel sin cerrar
    Verificar que wb.close() se llama siempre
    """

    def test_procesar_excel_cierra_workbook_en_exito(self, temp_dir):
        """Test que _procesar_excel cierra workbook en caso exitoso"""
        from sources.clientes_libres.help.transform_helpers import _procesar_excel
        from openpyxl import Workbook

        # Crear Excel de prueba
        excel_path = Path(temp_dir) / 'test.xlsx'
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'test'
        wb.save(excel_path)
        wb.close()

        mapping = {'campo1': {'cell': 'A1', 'type': 'str'}}
        sheet_names = ['Sheet']

        with patch('sources.clientes_libres.help.transform_helpers.load_workbook') as mock_load_wb:
            mock_wb = MagicMock()
            mock_wb.sheetnames = ['Sheet']
            mock_wb.__getitem__ = Mock(return_value=MagicMock())
            mock_load_wb.return_value = mock_wb

            # Ejecutar
            result = _procesar_excel(excel_path, mapping, sheet_names)

            # Verificar que se cerró el workbook
            assert mock_wb.close.called

    def test_procesar_excel_cierra_workbook_en_error(self, temp_dir):
        """Test que _procesar_excel cierra workbook incluso si hay error"""
        from sources.clientes_libres.help.transform_helpers import _procesar_excel

        excel_path = Path(temp_dir) / 'test.xlsx'

        mapping = {'campo1': {'cell': 'A1', 'type': 'str'}}
        sheet_names = ['Sheet']

        with patch('sources.clientes_libres.help.transform_helpers.load_workbook') as mock_load_wb:
            mock_wb = MagicMock()
            mock_wb.sheetnames = ['Sheet']
            # Hacer que falle al acceder a la hoja
            mock_wb.__getitem__.side_effect = Exception("Sheet access failed")
            mock_load_wb.return_value = mock_wb

            # Debe lanzar excepción
            with pytest.raises(Exception):
                _procesar_excel(excel_path, mapping, sheet_names)

            # Pero debe cerrar el workbook en finally
            assert mock_wb.close.called

    def test_procesar_excel_cierra_workbook_con_return_temprano(self, temp_dir):
        """Test que cierra workbook incluso con return temprano"""
        from sources.clientes_libres.help.transform_helpers import _procesar_excel

        excel_path = Path(temp_dir) / 'test.xlsx'

        mapping = {'campo1': {'cell': 'A1', 'type': 'str'}}
        sheet_names = ['HojaInexistente']  # Hoja que no existe

        with patch('sources.clientes_libres.help.transform_helpers.load_workbook') as mock_load_wb:
            mock_wb = MagicMock()
            mock_wb.sheetnames = ['Sheet']  # Hoja diferente
            mock_load_wb.return_value = mock_wb

            # Ejecutar (retorna temprano porque no hay hojas válidas)
            result = _procesar_excel(excel_path, mapping, sheet_names)

            # Debe cerrar el workbook incluso con return temprano
            assert mock_wb.close.called
            assert result.empty


# Ejecutar tests si se ejecuta directamente
if __name__ == '__main__':
    pytest.main([__file__, '-v'])
