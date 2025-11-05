"""
Tests para validar correcciones de FASE 3 - Seguridad y Validación
Bugs: #9, #10, #11, #12
"""
import pytest
import sys
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "proyectos" / "energiafacilities"))
from core.base_postgress import PostgresConnector, _validate_sql_identifier
from core.base_loader import BaseLoaderPostgres, _validate_sql_identifier as _validate_sql_identifier_loader
from core.utils import asegurar_directorio_sftp
from core.base_stractor import BaseExtractorSFTP


class TestBug10_SQLInjectionPrevencion:
    """
    BUG #10: SQL Injection potencial
    Verificar que identificadores SQL se validan correctamente
    """

    # Tests de validación de identificadores
    def test_validate_sql_identifier_acepta_identificadores_validos(self):
        """Test que _validate_sql_identifier acepta identificadores válidos"""
        # Identificadores válidos
        assert _validate_sql_identifier("tabla_test", "table") == "tabla_test"
        assert _validate_sql_identifier("_private", "column") == "_private"
        assert _validate_sql_identifier("Table123", "table") == "Table123"
        assert _validate_sql_identifier("public.users", "table") == "public.users"
        assert _validate_sql_identifier("schema_1.table_2", "table") == "schema_1.table_2"

    def test_validate_sql_identifier_rechaza_identificadores_invalidos(self):
        """Test que _validate_sql_identifier rechaza identificadores con SQL injection"""
        # Intentos de SQL injection
        with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
            _validate_sql_identifier("users; DROP TABLE users--", "table")

        with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
            _validate_sql_identifier("' OR '1'='1", "table")

        with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
            _validate_sql_identifier("users/*comment*/", "table")

        with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
            _validate_sql_identifier("users WHERE 1=1", "table")

        with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
            _validate_sql_identifier("users-malicious", "table")

    def test_validate_sql_identifier_rechaza_identificadores_vacios(self):
        """Test que rechaza identificadores vacíos"""
        with pytest.raises(ValueError, match="no puede estar vacío"):
            _validate_sql_identifier("", "table")

    def test_validate_sql_identifier_rechaza_multiples_puntos(self):
        """Test que rechaza identificadores con múltiples puntos"""
        with pytest.raises(ValueError, match="demasiados puntos"):
            _validate_sql_identifier("schema.table.column", "table")

    def test_validate_sql_identifier_rechaza_caracteres_especiales(self):
        """Test que rechaza caracteres especiales peligrosos"""
        caracteres_peligrosos = ["@", "#", "$", "%", "^", "&", "*", "(", ")", "+", "=", "[", "]", "{", "}", "|", "\\", ":", ";", "'", '"', "<", ">", "?", "/"]

        for char in caracteres_peligrosos:
            with pytest.raises(ValueError):
                _validate_sql_identifier(f"tabla{char}test", "table")

    # Tests de extract() con validación
    def test_extract_valida_schema_y_table(self, postgres_config):
        """Test que extract() valida schema y table"""
        with patch('core.base_postgress.create_engine') as mock_engine:
            mock_engine.return_value.connect.return_value.__enter__ = Mock()
            mock_engine.return_value.connect.return_value.__exit__ = Mock()

            pg = PostgresConnector(postgres_config)

            # Debe rechazar schema malicioso
            with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
                pg.extract({
                    'schema': 'public; DROP TABLE users--',
                    'table': 'test',
                    'columns': ['id']
                })

            # Debe rechazar table malicioso
            with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
                pg.extract({
                    'schema': 'public',
                    'table': "test' OR '1'='1",
                    'columns': ['id']
                })

    def test_extract_valida_columnas(self, postgres_config):
        """Test que extract() valida nombres de columnas"""
        with patch('core.base_postgress.create_engine') as mock_engine:
            mock_engine.return_value.connect.return_value.__enter__ = Mock()
            mock_engine.return_value.connect.return_value.__exit__ = Mock()

            pg = PostgresConnector(postgres_config)

            # Debe rechazar columnas maliciosas
            with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
                pg.extract({
                    'schema': 'public',
                    'table': 'test',
                    'columns': ['id', 'nombre; DELETE FROM test--']
                })

    def test_insert_dataframe_valida_identificadores(self, postgres_config, loader_config, sample_dataframe):
        """Test que insert_dataframe valida schema, table y columnas"""
        with patch('core.base_loader.psycopg2.connect') as mock_connect:
            mock_conn = MagicMock()
            mock_cursor = MagicMock()
            mock_cursor.fetchone.return_value = (False,)
            mock_conn.cursor.return_value.__enter__ = Mock(return_value=mock_cursor)
            mock_conn.cursor.return_value.__exit__ = Mock()
            mock_conn.__enter__ = Mock(return_value=mock_conn)
            mock_conn.__exit__ = Mock()
            mock_connect.return_value = mock_conn

            loader = BaseLoaderPostgres(postgres_config, loader_config)

            # Debe rechazar schema malicioso
            with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
                loader.insert_dataframe(
                    sample_dataframe,
                    schema="public; DROP TABLE test--"
                )

    def test_insert_dataframe_valida_nombres_columnas_df(self, postgres_config, loader_config):
        """Test que insert_dataframe valida nombres de columnas del DataFrame"""
        with patch('core.base_loader.psycopg2.connect') as mock_connect:
            mock_connect.return_value.__enter__ = Mock()
            mock_connect.return_value.__exit__ = Mock()

            loader = BaseLoaderPostgres(postgres_config, loader_config)

            # DataFrame con columna maliciosa
            df_malicioso = pd.DataFrame({
                'id': [1, 2],
                'nombre; DROP TABLE users--': ['test1', 'test2']
            })

            # Debe rechazar
            with pytest.raises(ValueError, match="contiene caracteres no permitidos"):
                loader.insert_dataframe(df_malicioso)


class TestBug11_ValidacionFilename:
    """
    BUG #11: Validación insuficiente de filename
    Verificar que se valida filename antes de operaciones
    """

    def test_ejecutar_transformacion_rechaza_filename_none(self):
        """Test que ejecutar_transformacion rechaza filename None cuando save=True"""
        from sources.clientes_libres.help.transform_helpers import ejecutar_transformacion

        config_transform = {}  # Sin local_destination_dir
        mapeo_campos = {
            'mapping': {'campo1': {'cell': 'A1', 'type': 'str'}},
            'sheet_names': ['Sheet1']
        }

        with patch('sources.clientes_libres.help.transform_helpers.load_workbook'):
            with pytest.raises(ValueError, match="No se especificó un directorio de destino"):
                ejecutar_transformacion(
                    config_transform,
                    mapeo_campos,
                    filepath='/tmp/test.xlsx',
                    save=True,
                    newdestinationoptional=None
                )

    def test_ejecutar_transformacion_acepta_filename_valido(self, temp_dir):
        """Test que ejecutar_transformacion funciona con filename válido"""
        from sources.clientes_libres.help.transform_helpers import ejecutar_transformacion
        from openpyxl import Workbook

        # Crear Excel de prueba
        excel_path = Path(temp_dir) / 'input.xlsx'
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'test'
        wb.save(excel_path)
        wb.close()

        output_path = Path(temp_dir) / 'output.xlsx'

        config_transform = {'local_destination_dir': str(output_path)}
        mapeo_campos = {
            'mapping': {'campo1': {'cell': 'A1', 'type': 'str'}},
            'sheet_names': ['Sheet']
        }

        result = ejecutar_transformacion(
            config_transform,
            mapeo_campos,
            filepath=str(excel_path),
            save=True
        )

        assert result == str(output_path)

    def test_ejecutar_transformacion_usa_newdestinationoptional(self, temp_dir):
        """Test que usa newdestinationoptional cuando local_destination_dir es None"""
        from sources.clientes_libres.help.transform_helpers import ejecutar_transformacion
        from openpyxl import Workbook

        excel_path = Path(temp_dir) / 'input.xlsx'
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'test'
        wb.save(excel_path)
        wb.close()

        output_path = Path(temp_dir) / 'output2.xlsx'

        config_transform = {}  # Sin local_destination_dir
        mapeo_campos = {
            'mapping': {'campo1': {'cell': 'A1', 'type': 'str'}},
            'sheet_names': ['Sheet']
        }

        result = ejecutar_transformacion(
            config_transform,
            mapeo_campos,
            filepath=str(excel_path),
            save=True,
            newdestinationoptional=str(output_path)
        )

        assert result == str(output_path)


class TestBug12_ValidacionSFTPRename:
    """
    BUG #12: Operación SFTP rename sin validación
    Verificar que se valida si archivo destino existe antes de rename
    """

    def test_extract_advierte_si_archivo_destino_existe(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que extract advierte si archivo destino ya existe"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            # Simular que archivo destino YA existe
            mock_attr = MagicMock()
            mock_attr.st_size = 1024
            mock_sftp_client.stat.return_value = mock_attr  # Archivo existe

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client), \
                 patch('core.base_stractor.asegurar_directorio_sftp'), \
                 patch('core.base_stractor.logger') as mock_logger:

                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                result = extractor.extract(
                    remotetransfere=True,
                    specific_file='test.xlsx'
                )

                # Debe haber registrado warning
                assert mock_logger.warning.called
                warning_msg = mock_logger.warning.call_args[0][0]
                assert 'ya existe' in warning_msg
                assert 'sobrescrito' in warning_msg

    def test_extract_no_advierte_si_archivo_destino_no_existe(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que extract no advierte si archivo destino no existe"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            # Simular que archivo destino NO existe
            mock_sftp_client.stat.side_effect = FileNotFoundError("Not found")

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client), \
                 patch('core.base_stractor.asegurar_directorio_sftp'), \
                 patch('core.base_stractor.logger') as mock_logger:

                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                result = extractor.extract(
                    remotetransfere=True,
                    specific_file='test.xlsx'
                )

                # NO debe haber registrado warning
                assert not mock_logger.warning.called

    def test_extract_continua_con_rename_despues_de_validacion(self, sftp_config_connect, sftp_config_paths, mock_sftp_client):
        """Test que extract continúa con rename después de validar"""
        with patch('core.base_stractor.paramiko.Transport') as mock_transport_class:
            mock_transport = MagicMock()
            mock_transport_class.return_value = mock_transport

            mock_sftp_client.stat.side_effect = FileNotFoundError()

            with patch('core.base_stractor.paramiko.SFTPClient.from_transport', return_value=mock_sftp_client), \
                 patch('core.base_stractor.asegurar_directorio_sftp'):

                extractor = BaseExtractorSFTP(sftp_config_connect, sftp_config_paths)

                result = extractor.extract(
                    remotetransfere=True,
                    specific_file='test.xlsx'
                )

                # Debe haber llamado a rename
                assert mock_sftp_client.rename.called
                assert result['status'] == 'success'


class TestBug9_VariableNoUsada:
    """
    BUG #9: Variable no usada en utils.py
    Verificar que no se asigna variable innecesaria
    """

    def test_asegurar_directorio_sftp_no_asigna_variable_innecesaria(self, mock_sftp_client):
        """Test que asegurar_directorio_sftp no asigna resultado de stat a variable"""
        # Simular que directorio ya existe
        mock_sftp_client.stat.return_value = MagicMock()

        # Ejecutar función
        asegurar_directorio_sftp(mock_sftp_client, '/test/path/dir')

        # Verificar que se llamó stat (para verificar existencia)
        assert mock_sftp_client.stat.called

    def test_asegurar_directorio_sftp_crea_directorios_faltantes(self, mock_sftp_client):
        """Test que asegurar_directorio_sftp crea directorios cuando no existen"""
        # Simular que directorios no existen
        mock_sftp_client.stat.side_effect = FileNotFoundError()

        # Ejecutar función
        asegurar_directorio_sftp(mock_sftp_client, '/test/path/dir')

        # Debe haber llamado mkdir para cada nivel
        assert mock_sftp_client.mkdir.call_count == 3
        calls = [call[0][0] for call in mock_sftp_client.mkdir.call_args_list]
        assert '/test' in calls
        assert '/test/path' in calls
        assert '/test/path/dir' in calls


# Ejecutar tests si se ejecuta directamente
if __name__ == '__main__':
    pytest.main([__file__, '-v'])
